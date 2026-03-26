"""
Parse Philippine survey XLSM files (TD Check workbooks).

These workbooks have standardized sheets:
  - Title Data: lot metadata (name, title, area, owner, location)
  - E2C: bearing-distance traverse data
  - Correction Comparison: flagged corrections
  - Commonline Checker: shared boundary validation

Cell positions are fixed across all files (same template).
"""
import openpyxl


def parse_title_data(file_path: str) -> dict:
    """
    Parse the Title Data sheet from an XLSM file.

    Cell positions (fixed template):
      A1/B1: DECREE/TCT NO
      A2/B2: LOT NO
      A5/B5: SURVEY NUMBER
      A6/B6: DATE OF ORIGINAL SURVEY (may be empty)
      A7/B7: BARRIO/BARANGAY
      A8/B8: MUN/CITY
      A9/B9: PROVINCE
      A10/B10: ISLAND
      A11/B11: REGISTERED OWNER
      A12/B12: TIE POINT
      A13/B13: AREA (numeric)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb["Title Data"]

    def cell(col, row):
        val = ws[f"{col}{row}"].value
        return str(val).strip() if val is not None else ""

    area_raw = ws["B13"].value
    stated_area = float(area_raw) if area_raw is not None else 0.0

    result = {
        "title_number": cell("B", 1),
        "lot_name": cell("B", 2),
        "survey_number": cell("B", 5),
        "date_of_original_survey": cell("B", 6),
        "barangay": cell("B", 7),
        "municipality": cell("B", 8),
        "province": cell("B", 9),
        "island": cell("B", 10),
        "owner": cell("B", 11),
        "tie_point": cell("B", 12),
        "stated_area": stated_area,
    }

    wb.close()
    return result


def parse_e2c(file_path: str) -> dict:
    """
    Parse the E2C sheet (bearing-distance traverse) from an XLSM file.

    Cell positions (fixed template):
      Config:
        B15: decimal places, B16: scale, B17: CIRCLE(Y/N),
        B18: ADJ(Y/N), B19: BEARING(Y/N), B23: PS/DENR-LRA
      Tie line (row 19):
        P19: N/S, Q19: E/W, R19: degrees, S19: minutes, T19: distance
      Lot name: L25 (column 12)
      Corner 1 coordinates: BG25 (col 59), BH25 (col 60)
      Traverse rows 26+:
        A: adjacent lot name (sticky), B: adjacent survey number,
        I: from corner, K: to corner,
        P: N/S, Q: E/W, R: degrees, S: minutes, T: distance,
        BG: northing, BH: easting
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    ws = wb["E2C"]

    def cell(row, col):
        v = ws.cell(row, col).value
        return v

    def cell_str(row, col):
        v = ws.cell(row, col).value
        return str(v).strip() if v is not None else ""

    def parse_int(val):
        if val is None:
            return None
        s = str(val).strip()
        if not s:
            return None
        # Strip non-digit chars (degree symbols, etc.)
        digits = ''.join(c for c in s if c.isdigit())
        return int(digits) if digits else None

    def parse_bearing_row(row):
        ns = cell_str(row, 16)  # P
        ew = cell_str(row, 17)  # Q
        deg = parse_int(cell(row, 18))  # R
        mins = parse_int(cell(row, 19))  # S
        dist_raw = cell(row, 20)  # T
        dist = float(dist_raw) if dist_raw is not None else 0.0
        return {
            "bearing": {"ns": ns, "ew": ew, "degrees": deg or 0, "minutes": mins or 0},
            "distance": dist,
        }

    # Config
    config = {
        "decimal_places": parse_int(cell(15, 2)) or 3,
        "scale": parse_int(cell(16, 2)) or 1000,
        "circle": cell_str(17, 2).upper(),
        "adj": cell_str(18, 2).upper(),
        "bearing": cell_str(19, 2).upper(),
        "authority": cell_str(23, 2),
    }

    # Tie line (row 19)
    tie_data = parse_bearing_row(19)
    tie_line = tie_data if tie_data["distance"] > 0 else None

    # Lot name (L25 = column 12)
    lot_name = cell_str(25, 12)

    # Corner 1 coordinates from row 25
    computed_coordinates = []
    bg25 = cell(25, 59)
    bh25 = cell(25, 60)
    if bg25 is not None and bh25 is not None:
        computed_coordinates.append({
            "corner": 1,
            "northing": float(bg25),
            "easting": float(bh25),
        })

    # Traverse rows starting at 26
    lines = []
    last_adj_lot = ""
    last_adj_survey = ""
    row = 26
    while True:
        from_corner_raw = cell(row, 9)  # I
        if from_corner_raw is None:
            break
        from_corner = parse_int(from_corner_raw)
        if from_corner is None:
            break

        to_corner = parse_int(cell(row, 11)) or 0  # K

        # Sticky adjacent lot/survey
        adj_lot = cell_str(row, 1)  # A
        adj_survey = cell_str(row, 2)  # B
        if adj_lot:
            last_adj_lot = adj_lot
        if adj_survey:
            last_adj_survey = adj_survey

        bearing_data = parse_bearing_row(row)

        lines.append({
            "from_corner": from_corner,
            "to_corner": to_corner,
            "bearing": bearing_data["bearing"],
            "distance": bearing_data["distance"],
            "adjacent_lot": last_adj_lot,
            "adjacent_survey": last_adj_survey,
        })

        # Computed coordinates for the endpoint
        bg = cell(row, 59)
        bh = cell(row, 60)
        if bg is not None and bh is not None:
            computed_coordinates.append({
                "corner": to_corner,
                "northing": float(bg),
                "easting": float(bh),
            })

        row += 1

    wb.close()
    return {
        "config": config,
        "tie_line": tie_line,
        "lot_name": lot_name,
        "lines": lines,
        "computed_coordinates": computed_coordinates,
    }
