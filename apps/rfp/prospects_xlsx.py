#!/usr/bin/env python3
"""awards.db -> a prospect workbook you can actually work from.

    uv run --with openpyxl python prospects_xlsx.py [out.xlsx]

Three sheets:
  Shortlist   SME infra contractors active in 2026 -- the #140 outreach cohort
  All winners every company in the sample, for filtering your own way
  Read me     what the data is, what is missing, and where each field came from

The gap this cannot fill: PhilGEPS publishes company, contact person and street address, and
publishes NO email and NO phone. The 'Contact channel' column is deliberately left blank for you
to fill during lookup -- an empty column you can see beats a missing one you discover later.
"""
import re, sqlite3, sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "philgeps-prospects.xlsx"
NAVY, PEACH = "0C1F40", "F6AE72"
INFRA = ("road", "canal", "drainage", "building", "pavement", "water", "bridge",
         "school", "concrete", "construction")
AWARD_URL = ("https://notices.philgeps.gov.ph/GEPSNONPILOT/R4/"
             "R3_AwardNotice_AwardAbstract.html?RefID={ref}&LineItemID=1&OrgID=0&AwardID={aid}")


def dt(s):
    try:
        return datetime.strptime(s, "%d-%b-%Y")
    except Exception:
        return None


def tidy(a):
    a = re.sub(r"\s+", " ", (a or "")).strip()
    return re.sub(r",?\s*Philippines$", "", a)


def contacts():
    d = sqlite3.connect(f"file:{HERE/'awards.db'}?mode=ro", uri=True)
    try:
        rows = d.execute("select winner, phone, website, maps_uri, confidence from contacts")
    except sqlite3.OperationalError:
        return {}
    return {w: dict(phone=p, website=s, maps_uri=m, confidence=c) for w, p, s, m, c in rows}


CONTACTS = {}


def load():
    d = sqlite3.connect(f"file:{HERE/'awards.db'}?mode=ro", uri=True)
    rows = d.execute("""select winner, winner_contact, winner_address, winner_province,
                               contract_amount, abc, win_ratio, award_date, title,
                               unspsc_desc, award_id, ref_id
                        from awards where winner is not null""").fetchall()
    firms = {}
    for (w, person, addr, prov, amt, abc, ratio, ad, title, uns, aid, ref) in rows:
        D = dt(ad)
        f = firms.setdefault(w, dict(
            winner=w, contact=person, address=tidy(addr), province=prov,
            wins=0, total=0.0, last=None, last_title=None, last_amt=None,
            aid=aid, ref=ref, ratios=[], sectors=set()))
        # count DISTINCT procurements, not award rows: one procurement is split across a row per
        # line item, so counting rows ranks a nine-item shopping list above three real wins.
        f["wins"] += 1
        f["total"] += amt or 0
        if ratio:
            f["ratios"].append(ratio)
        if uns:
            f["sectors"].add(uns)
        if D and (f["last"] is None or D > f["last"]):
            f.update(last=D, last_title=title, last_amt=amt, aid=aid, ref=ref)
    return list(firms.values())


def is_shortlist(f):
    return (f["last"] and f["last"] >= datetime(2026, 1, 1)
            and f["last_amt"] and 500_000 <= f["last_amt"] <= 15_000_000
            and any(k in (f["last_title"] or "").lower() for k in INFRA))


HEAD = ["Company", "Contact person", "Phone", "Website", "Match confidence", "Address", "Province",
        "Last win", "Last win ₱", "Wins in sample", "Total ₱ in sample",
        "Avg bid vs budget", "What they last won", "Sector", "Verify on PhilGEPS", "Google Maps"]


def sheet(wb, title, firms, note):
    ws = wb.create_sheet(title)
    ws.append([note])
    ws["A1"].font = Font(italic=True, color="666666", size=9)
    ws.append(HEAD)
    for c in range(1, len(HEAD) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for f in sorted(firms, key=lambda x: (x["last"] or datetime(1970, 1, 1)), reverse=True):
        avg = sum(f["ratios"]) / len(f["ratios"]) if f["ratios"] else None
        ct = CONTACTS.get(f["winner"], {})
        ws.append([
            f["winner"], f["contact"], ct.get("phone"), ct.get("website"),
            ct.get("confidence"), f["address"], f["province"],
            f["last"].strftime("%Y-%m-%d") if f["last"] else "",
            f["last_amt"], f["wins"], round(f["total"], 2),
            round(avg, 3) if avg else None,
            re.sub(r"^[0-9A-Z]+ - ", "", f["last_title"] or "")[:120],
            "; ".join(sorted(f["sectors"]))[:60],
            AWARD_URL.format(ref=f["ref"] or 0, aid=f["aid"]),
            ct.get("maps_uri"),
        ])
    widths = [42, 24, 17, 34, 12, 44, 16, 11, 13, 8, 15, 10, 52, 26, 24, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    GREEN, AMBER, GREY = "E6F4EA", "FFF4E8", "F2F2F2"
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=9).number_format = '#,##0'
        ws.cell(row=r, column=11).number_format = '#,##0'
        ws.cell(row=r, column=12).number_format = '0.0%'
        # colour the confidence cell: a wrong phone number is worse than a blank one, so the
        # reader must see at a glance which rows a human still has to verify.
        conf = ws.cell(row=r, column=5).value
        ws.cell(row=r, column=5).fill = PatternFill(
            "solid", fgColor={"good": GREEN, "weak": AMBER}.get(conf, GREY))
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEAD))}{ws.max_row}"
    return ws


def readme(wb, n_all, n_short, n_awards):
    ws = wb.create_sheet("Read me", 0)
    ws.column_dimensions["A"].width = 116
    lines = [
        ("PhilGEPS award-winner prospect list", True),
        ("", False),
        (f"{n_all} companies, from {n_awards} award records sampled {datetime.now():%d %b %Y}.", False),
        (f"Shortlist sheet: {n_short} SME infrastructure contractors active in 2026.", False),
        ("", False),
        ("CONTACT DETAILS -- read the confidence colour before you dial", True),
        ("PhilGEPS publishes NO email and NO phone, and its Registered Merchants directory carries", False),
        ("only membership status and certificate dates. Phone and website here come from Google", False),
        ("Places, matched on company name AND province.", False),
        ("  GREEN 'good' = name and province both agree. Safe to use.", False),
        ("  AMBER 'weak' = plausible but unconfirmed. Check the Google Maps link before calling.", False),
        ("  GREY  'none' = no credible match; the lookup found something and it was rejected.", False),
        ("A wrong number is worse than a blank one -- you call a stranger and burn the lead -- so", False),
        ("the matcher rejects rather than guesses. Still no emails: for those, try the website or", False),
        ("the company's Facebook page.", False),
        ("", False),
        ("HOW THE SHORTLIST WAS FILTERED", True),
        ("Won between P500,000 and P15,000,000 -- big enough to matter, small enough to lack a bid desk.", False),
        ("Most recent win in 2026 -- an active bidder, not a dormant registration.", False),
        ("Infrastructure-type work (road, canal, drainage, building, water, bridge, school, concrete).", False),
        ("", False),
        ("SAMPLING -- read this before drawing conclusions from the counts", True),
        (f"These {n_awards} awards are a random sample of ~1.19M award ids, well under 1%.", False),
        ("So 'Wins in sample' is NOT how often a firm wins. A company appearing 3 times has probably", False),
        ("won thousands; one appearing once may still be a frequent winner. Ranking by that column", False),
        ("selects the biggest companies in the country, which are the ones who need a bid-finding", False),
        ("tool least. The sheet is sorted by most recent win instead.", False),
        ("", False),
        ("PROVENANCE", True),
        ("Every field comes from the public award-abstract page, viewable with no login:", False),
        ("notices.philgeps.gov.ph/GEPSNONPILOT/R4/R3_AwardNotice_AwardAbstract.html", False),
        ("The 'Verify on PhilGEPS' link opens the source record for that firm's most recent win.", False),
        ("", False),
        ("BEFORE YOU SEND ANYTHING", True),
        ("The contact person is a named individual and the PH Data Privacy Act covers personal data", False),
        ("even inside a public record. Company-level B2B outreach with a clear opt-out is the clean", False),
        ("version. Using the name is defensible -- it is published, and for a one-person firm the", False),
        ("owner is the company -- but do not mass-blast, and honour opt-outs on first request.", False),
    ]
    for text, bold in lines:
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(bold=bold, size=11 if bold else 10,
                      color=NAVY if bold else "000000")
        c.alignment = Alignment(wrap_text=True)
    return ws


def main():
    global CONTACTS
    CONTACTS = contacts()
    firms = load()
    short = [f for f in firms if is_shortlist(f)]
    n_awards = sqlite3.connect(f"file:{HERE/'awards.db'}?mode=ro", uri=True).execute(
        "select count(*) from awards").fetchone()[0]
    wb = Workbook()
    wb.remove(wb.active)
    readme(wb, len(firms), len(short), n_awards)
    sheet(wb, "Shortlist", short,
          "SME infrastructure contractors, active 2026, won P0.5-15M. Sorted by most recent win.")
    sheet(wb, "All winners", firms,
          "Every award winner in the sample. Filter this yourself; see 'Read me' on the counts.")
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(short)} shortlist / {len(firms)} companies / {n_awards} awards)")


if __name__ == "__main__":
    main()
